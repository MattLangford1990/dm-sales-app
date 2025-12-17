from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import Response, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from datetime import datetime, timedelta
from jose import JWTError, jwt
from typing import Optional, List, Dict
import re
import os
import io

import json
from config import get_settings
from agents import get_agent, get_agent_brands, verify_agent_pin, list_agents, get_all_brand_patterns, is_admin, list_all_agents_admin, create_agent, update_agent, delete_agent, get_all_brands, change_agent_pin
import zoho_api

# Load pack quantities
PACK_QUANTITIES_FILE = os.path.join(os.path.dirname(__file__), "pack_quantities.json")
def load_pack_quantities():
    if os.path.exists(PACK_QUANTITIES_FILE):
        with open(PACK_QUANTITIES_FILE, "r") as f:
            return json.load(f)
    return {}

_pack_quantities = load_pack_quantities()

# Load image URLs (Cloudinary URLs for Elvang etc)
IMAGE_URLS_FILE = os.path.join(os.path.dirname(__file__), "image_urls.json")
def load_image_urls():
    if os.path.exists(IMAGE_URLS_FILE):
        with open(IMAGE_URLS_FILE, "r") as f:
            urls = json.load(f)
            print(f"STARTUP: Loaded {len(urls)} image URLs from image_urls.json")
            return urls
    print("STARTUP: No image_urls.json found")
    return {}

_image_urls = load_image_urls()

settings = get_settings()
security = HTTPBearer()

app = FastAPI(
    title="DM Sales App API",
    description="Sales order management for DM Brands agents",
    version="1.0.0"
)

# CORS for the frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, restrict to your domain
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============ Pydantic Models ============

class LoginRequest(BaseModel):
    agent_id: str
    pin: str


class LoginResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    agent_name: str
    brands: List[str]


class TokenData(BaseModel):
    agent_id: str
    agent_name: str
    brands: List[str]


class CustomerCreate(BaseModel):
    company_name: str
    contact_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    billing_address: Optional[str] = None
    shipping_address: Optional[str] = None
    booking_requirements: Optional[str] = None
    payment_terms: Optional[str] = None


class OrderLineItem(BaseModel):
    item_id: str
    name: str
    sku: str
    quantity: int
    rate: float
    discount_percent: float = 0


class OrderCreate(BaseModel):
    customer_id: str
    customer_name: str
    line_items: List[OrderLineItem]
    notes: Optional[str] = None
    reference_number: Optional[str] = None
    delivery_date: Optional[str] = None
    delivery_charge: Optional[float] = 0


# ============ Auth Helpers ============

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=settings.access_token_expire_minutes)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, settings.secret_key, algorithm=settings.algorithm)


async def get_current_agent(credentials: HTTPAuthorizationCredentials = Depends(security)) -> TokenData:
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Invalid authentication credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(credentials.credentials, settings.secret_key, algorithms=[settings.algorithm])
        agent_id: str = payload.get("sub")
        if agent_id is None:
            raise credentials_exception
        return TokenData(
            agent_id=agent_id,
            agent_name=payload.get("agent_name", ""),
            brands=payload.get("brands", [])
        )
    except JWTError:
        raise credentials_exception


# ============ Auth Routes ============

@app.get("/api/debug/raw-items")
async def debug_raw_items():
    """Debug: Show raw items from Zoho (no auth required)"""
    try:
        # Fetch 500 items to get a good sample of brands
        all_items = []
        for page in range(1, 6):  # 5 pages of 100 = 500 items
            response = await zoho_api.get_items(page=page, per_page=100)
            items = response.get("items", [])
            all_items.extend(items)
            if len(items) < 100:
                break
        
        # Collect unique brand values
        brands = set()
        manufacturers = set()
        for item in all_items:
            if item.get("brand"):
                brands.add(item.get("brand"))
            if item.get("manufacturer"):
                manufacturers.add(item.get("manufacturer"))
        
        return {
            "total_items_fetched": len(all_items),
            "unique_brands": sorted(list(brands)),
            "unique_manufacturers": sorted(list(manufacturers)),
            "sample_items": [
                {
                    "name": item.get("name"),
                    "sku": item.get("sku"),
                    "brand": item.get("brand"),
                    "manufacturer": item.get("manufacturer")
                }
                for item in all_items[:10]
            ]
        }
    except Exception as e:
        import traceback
        return {"error": str(e), "traceback": traceback.format_exc()}

@app.get("/api/debug/find-brand/{brand_name}")
async def debug_find_brand(brand_name: str):
    """Debug: Search for items containing a brand name"""
    try:
        all_items = []
        for page in range(1, 11):  # 10 pages
            response = await zoho_api.get_items(page=page, per_page=100)
            items = response.get("items", [])
            all_items.extend(items)
            if len(items) < 100:
                break
        
        # Find items matching the brand
        matching = []
        brand_lower = brand_name.lower()
        for item in all_items:
            item_brand = (item.get("brand") or "").lower()
            item_manufacturer = (item.get("manufacturer") or "").lower()
            item_name = (item.get("name") or "").lower()
            
            if brand_lower in item_brand or brand_lower in item_manufacturer or brand_lower in item_name:
                matching.append({
                    "name": item.get("name"),
                    "sku": item.get("sku"),
                    "brand": item.get("brand"),
                    "manufacturer": item.get("manufacturer")
                })
        
        return {
            "search_term": brand_name,
            "total_items_searched": len(all_items),
            "matches_found": len(matching),
            "matching_items": matching[:20]
        }
    except Exception as e:
        import traceback
        return {"error": str(e), "traceback": traceback.format_exc()}

@app.get("/api/debug/feed-check/{brand_name}")
async def debug_feed_check(brand_name: str):
    """Debug: Check feed in database for a specific brand (no auth required)"""
    try:
        from database import SessionLocal, ProductFeed
        db = SessionLocal()
        feed = db.query(ProductFeed).filter(ProductFeed.id == "main").first()
        db.close()
        
        if not feed or not feed.feed_json:
            return {"error": "No feed in database"}
        
        data = json.loads(feed.feed_json)
        products = data.get("products", [])
        
        # Find products matching brand
        brand_lower = brand_name.lower()
        matching = [p for p in products if brand_lower in (p.get("brand") or "").lower()]
        with_images = [p for p in matching if p.get("image_url")]
        
        return {
            "brand": brand_name,
            "feed_generated_at": data.get("generated_at"),
            "total_products_in_feed": len(products),
            "matching_brand": len(matching),
            "with_image_url": len(with_images),
            "image_urls_loaded": len(_image_urls),
            "sample_products": [
                {
                    "sku": p.get("sku"),
                    "name": p.get("name"),
                    "image_url": p.get("image_url", "")[:80] + "..." if p.get("image_url") else None
                }
                for p in matching[:5]
            ]
        }
    except Exception as e:
        import traceback
        return {"error": str(e), "traceback": traceback.format_exc()}

@app.get("/api/debug/live-products/{brand_name}")
async def debug_live_products(brand_name: str):
    """Debug: Simulate what /api/products returns for a brand (no auth)"""
    try:
        all_zoho_items = await zoho_api.get_all_items_cached()
        items = filter_items_by_brand(all_zoho_items, [brand_name])
        items = [i for i in items if i.get("status") != "inactive"][:5]
        
        products = []
        for item in items:
            sku = item.get("sku", "")
            img_url = _image_urls.get(sku) or item.get("image_url")
            if img_url and not img_url.startswith('http'):
                img_url = None
            products.append({
                "sku": sku,
                "name": item.get("name"),
                "image_url": img_url,
                "image_url_type": type(img_url).__name__,
                "raw_image_url": item.get("image_url"),
                "from_image_urls_json": _image_urls.get(sku)
            })
        
        return {
            "brand": brand_name,
            "products": products
        }
    except Exception as e:
        import traceback
        return {"error": str(e), "traceback": traceback.format_exc()}

@app.get("/api/debug/search/{search_term}")
async def debug_search(search_term: str):
    """Debug: Use Zoho's search API directly"""
    try:
        response = await zoho_api.get_items(page=1, per_page=100, search=search_term)
        items = response.get("items", [])
        
        return {
            "search_term": search_term,
            "results_count": len(items),
            "items": [
                {
                    "name": item.get("name"),
                    "sku": item.get("sku"),
                    "brand": item.get("brand"),
                    "manufacturer": item.get("manufacturer"),
                    "status": item.get("status")
                }
                for item in items[:20]
            ]
        }
    except Exception as e:
        import traceback
        return {"error": str(e), "traceback": traceback.format_exc()}

@app.get("/api/debug/products-flow")
async def debug_products_flow():
    """Debug: Show exactly what the products endpoint does for Kate's brands"""
    try:
        # Simulate Kate's brands
        kate_brands = ["Remember", "Räder", "My Flame", "Ideas4Seasons"]
        brand_search_terms = get_all_brand_patterns(kate_brands)
        
        results = {
            "kate_brands": kate_brands,
            "brand_search_terms": brand_search_terms,
            "search_results": {}
        }
        
        seen_ids = set()
        all_items = []
        
        for brand_term in brand_search_terms:
            response = await zoho_api.get_items(page=1, per_page=50, search=brand_term)
            brand_items = response.get("items", [])
            
            results["search_results"][brand_term] = {
                "count": len(brand_items),
                "sample": [item.get("name") for item in brand_items[:3]]
            }
            
            for item in brand_items:
                item_id = item.get("item_id")
                if item_id not in seen_ids:
                    seen_ids.add(item_id)
                    all_items.append(item)
        
        # Filter
        filtered = filter_items_by_brand(all_items, kate_brands)
        
        results["total_unique_items"] = len(all_items)
        results["after_filter"] = len(filtered)
        results["sample_filtered"] = [item.get("name") for item in filtered[:10]]
        
        return results
    except Exception as e:
        import traceback
        return {"error": str(e), "traceback": traceback.format_exc()}

@app.get("/api/debug/brands")
async def debug_brands(agent: TokenData = Depends(get_current_agent)):
    """Debug: Show unique brand values from Zoho items"""
    try:
        response = await zoho_api.get_items(page=1, per_page=200)
        items = response.get("items", [])
        
        # Collect all unique values from brand-related fields
        brands_found = set()
        sample_items = []
        
        for item in items[:50]:  # First 50 items
            # Get ALL keys from the item to see what fields exist
            all_keys = list(item.keys())
            
            brand = item.get("brand") or ""
            manufacturer = item.get("manufacturer") or ""
            cf_brand = item.get("cf_brand") or ""
            group_name = item.get("group_name") or ""
            category_name = item.get("category_name") or ""
            
            if brand: brands_found.add(f"brand: {brand}")
            if manufacturer: brands_found.add(f"manufacturer: {manufacturer}")
            if cf_brand: brands_found.add(f"cf_brand: {cf_brand}")
            if group_name: brands_found.add(f"group_name: {group_name}")
            if category_name: brands_found.add(f"category_name: {category_name}")
            
            sample_items.append({
                "name": item.get("name"),
                "sku": item.get("sku"),
                "all_keys": all_keys,
                "brand": brand,
                "manufacturer": manufacturer,
                "cf_brand": cf_brand,
                "group_name": group_name,
                "category_name": category_name,
                "raw_item": item  # Include full raw data
            })
        
        return {
            "agent_brands": agent.brands,
            "total_items": len(items),
            "unique_values_found": sorted(list(brands_found)),
            "sample_items": sample_items[:5]  # Just 5 items with full data
        }
    except Exception as e:
        import traceback
        return {"error": str(e), "traceback": traceback.format_exc()}

@app.post("/api/auth/login", response_model=LoginResponse)
async def login(request: LoginRequest):
    """Agent login with ID and PIN"""
    if not verify_agent_pin(request.agent_id, request.pin):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid agent ID or PIN"
        )
    
    agent = get_agent(request.agent_id)
    brands = get_agent_brands(request.agent_id)
    
    token = create_access_token({
        "sub": request.agent_id,
        "agent_name": agent["name"],
        "brands": brands
    })
    
    return LoginResponse(
        access_token=token,
        agent_name=agent["name"],
        brands=brands
    )


@app.get("/api/auth/me")
async def get_me(agent: TokenData = Depends(get_current_agent)):
    """Get current agent info"""
    return {
        "agent_id": agent.agent_id,
        "agent_name": agent.agent_name,
        "brands": agent.brands,
        "is_admin": is_admin(agent.agent_id)
    }


class ChangePinRequest(BaseModel):
    current_pin: str
    new_pin: str


@app.post("/api/auth/change-pin")
async def change_pin(request: ChangePinRequest, agent: TokenData = Depends(get_current_agent)):
    """Change the current agent's PIN"""
    try:
        change_agent_pin(agent.agent_id, request.current_pin, request.new_pin)
        return {"message": "PIN changed successfully"}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# ============ Admin Routes ============

class AgentCreate(BaseModel):
    agent_id: str
    name: str
    pin: str
    brands: List[str]
    commission_rate: float = 0.125

class AgentUpdate(BaseModel):
    name: Optional[str] = None
    pin: Optional[str] = None
    brands: Optional[List[str]] = None
    commission_rate: Optional[float] = None
    active: Optional[bool] = None


def require_admin(agent: TokenData = Depends(get_current_agent)) -> TokenData:
    """Dependency that requires admin access"""
    if not is_admin(agent.agent_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required"
        )
    return agent


@app.get("/api/admin/agents")
async def admin_list_agents(agent: TokenData = Depends(require_admin)):
    """List all agents with full details (admin only)"""
    return {
        "agents": list_all_agents_admin(),
        "available_brands": get_all_brands()
    }


@app.post("/api/admin/agents")
async def admin_create_agent(
    new_agent: AgentCreate,
    agent: TokenData = Depends(require_admin)
):
    """Create a new agent (admin only)"""
    try:
        created = create_agent(
            agent_id=new_agent.agent_id,
            name=new_agent.name,
            pin=new_agent.pin,
            brands=new_agent.brands,
            commission_rate=new_agent.commission_rate
        )
        return {"message": "Agent created successfully", "agent_id": new_agent.agent_id}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.put("/api/admin/agents/{agent_id}")
async def admin_update_agent(
    agent_id: str,
    updates: AgentUpdate,
    agent: TokenData = Depends(require_admin)
):
    """Update an agent (admin only)"""
    try:
        # Convert to dict, excluding None values
        update_dict = {k: v for k, v in updates.dict().items() if v is not None}
        updated = update_agent(agent_id, update_dict)
        return {"message": "Agent updated successfully", "agent_id": agent_id}
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.delete("/api/admin/agents/{agent_id}")
async def admin_delete_agent(
    agent_id: str,
    agent: TokenData = Depends(require_admin)
):
    """Delete or deactivate an agent (admin only)"""
    try:
        delete_agent(agent_id)
        return {"message": "Agent deleted successfully", "agent_id": agent_id}
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.post("/api/cron/cleanup-brands")
async def admin_cleanup_brands(secret: str = None):
    """Remove Elvang and GEFU from all agents' brands"""
    if secret != "dmbrands_cleanup_2025":
        raise HTTPException(status_code=403, detail="Invalid secret")
    
    from database import SessionLocal
    from agents import AgentModel
    
    db = SessionLocal()
    try:
        agents = db.query(AgentModel).all()
        updated = 0
        for ag in agents:
            if ag.brands:
                new_brands = [b for b in ag.brands if b not in ["Elvang", "GEFU"]]
                if len(new_brands) != len(ag.brands):
                    ag.brands = new_brands
                    updated += 1
        db.commit()
        return {"message": f"Updated {updated} agents", "removed": ["Elvang", "GEFU"]}
    finally:
        db.close()


@app.post("/api/admin/refresh-cache")
async def admin_refresh_cache(agent: TokenData = Depends(require_admin)):
    """Force refresh of the products cache (admin only)"""
    zoho_api.invalidate_items_cache()
    # Trigger a fresh fetch
    items = await zoho_api.get_all_items_cached()
    return {
        "message": "Cache refreshed successfully",
        "total_items": len(items),
        "refreshed_at": datetime.utcnow().isoformat()
    }


class CloudinarySyncRequest(BaseModel):
    dry_run: bool = True  # Default to dry run for safety
    limit: Optional[int] = None  # Limit number of images to process (for testing)


@app.post("/api/admin/sync-images-to-cloudinary")
async def admin_sync_images_to_cloudinary(
    request: CloudinarySyncRequest,
    agent: TokenData = Depends(require_admin)
):
    """
    Sync all product images from Zoho to Cloudinary (admin only).
    
    - dry_run=True: Just report what would be synced (default)
    - dry_run=False: Actually upload missing images to Cloudinary
    - limit: Process only first N products (for testing)
    """
    import httpx
    
    # Check Cloudinary credentials
    if not settings.cloudinary_cloud_name or not settings.cloudinary_api_key:
        raise HTTPException(
            status_code=500, 
            detail="Cloudinary credentials not configured. Add CLOUDINARY_CLOUD_NAME, CLOUDINARY_API_KEY, CLOUDINARY_API_SECRET to .env"
        )
    
    results = {
        "dry_run": request.dry_run,
        "total_products": 0,
        "with_sku": 0,
        "already_in_cloudinary": 0,
        "missing_in_cloudinary": 0,
        "uploaded": 0,
        "no_image_in_zoho": 0,
        "upload_failed": 0,
        "errors": [],
        "uploaded_skus": [],
        "missing_skus": []  # SKUs that need uploading
    }
    
    # Get all products from Zoho cache
    all_items = await zoho_api.get_all_items_cached()
    results["total_products"] = len(all_items)
    
    # Filter to items with SKU
    items_with_sku = [item for item in all_items if item.get("sku")]
    results["with_sku"] = len(items_with_sku)
    
    # Apply limit if specified
    if request.limit:
        items_with_sku = items_with_sku[:request.limit]
    
    # Cloudinary check URL pattern
    cloudinary_base = f"https://res.cloudinary.com/{settings.cloudinary_cloud_name}/image/upload"
    
    async with httpx.AsyncClient(timeout=30.0) as client:
        for item in items_with_sku:
            sku = item.get("sku")
            
            # Check if image exists in Cloudinary
            cloudinary_url = f"{cloudinary_base}/products/{sku}.jpg"
            try:
                check_response = await client.head(cloudinary_url)
                if check_response.status_code == 200:
                    results["already_in_cloudinary"] += 1
                    continue
            except:
                pass  # Assume not in Cloudinary if check fails
            
            # Image not in Cloudinary - check if we can get it from Zoho
            results["missing_in_cloudinary"] += 1
            results["missing_skus"].append(sku)
            
            if request.dry_run:
                continue  # Don't actually upload in dry run
            
            # Try to get image from Zoho
            try:
                image_data = await zoho_api.get_item_image(item.get("item_id"))
                
                if not image_data:
                    results["no_image_in_zoho"] += 1
                    continue
                
                # Upload to Cloudinary
                upload_url = f"https://api.cloudinary.com/v1_1/{settings.cloudinary_cloud_name}/image/upload"
                
                # Cloudinary upload with authentication
                import hashlib
                import time
                
                timestamp = str(int(time.time()))
                public_id = f"products/{sku}"
                
                # Generate signature
                params_to_sign = f"public_id={public_id}&timestamp={timestamp}{settings.cloudinary_api_secret}"
                signature = hashlib.sha1(params_to_sign.encode()).hexdigest()
                
                # Upload
                upload_response = await client.post(
                    upload_url,
                    data={
                        "public_id": public_id,
                        "timestamp": timestamp,
                        "api_key": settings.cloudinary_api_key,
                        "signature": signature,
                        "overwrite": "true"
                    },
                    files={"file": (f"{sku}.jpg", image_data, "image/jpeg")}
                )
                
                if upload_response.status_code == 200:
                    results["uploaded"] += 1
                    results["uploaded_skus"].append(sku)
                else:
                    results["upload_failed"] += 1
                    results["errors"].append({
                        "sku": sku,
                        "error": f"Upload failed: {upload_response.status_code} - {upload_response.text[:200]}"
                    })
                    
            except Exception as e:
                results["upload_failed"] += 1
                results["errors"].append({"sku": sku, "error": str(e)})
    
    # Limit the lists in response to avoid huge payloads
    results["missing_skus"] = results["missing_skus"][:100]  # First 100 missing
    results["errors"] = results["errors"][:50]  # First 50 errors
    
    return results


@app.get("/api/admin/stats")
async def admin_get_stats(agent: TokenData = Depends(require_admin)):
    """Get admin dashboard stats (admin only)"""
    try:
        # Get recent orders for stats
        response = await zoho_api.get_sales_orders(page=1)
        orders = response.get("salesorders", [])
        
        # Calculate stats
        today = datetime.now().strftime("%Y-%m-%d")
        orders_today = [o for o in orders if o.get("date") == today]
        
        # Get orders by agent (from notes)
        agents_list = list_all_agents_admin()
        orders_by_agent = {}
        for a in agents_list:
            agent_name = a["name"]
            count = len([o for o in orders if f"Order placed by {agent_name}" in (o.get("notes") or "")])
            if count > 0:
                orders_by_agent[agent_name] = count
        
        return {
            "total_agents": len(agents_list),
            "active_agents": len([a for a in agents_list if a.get("active", True)]),
            "orders_today": len(orders_today),
            "orders_today_value": sum(o.get("total", 0) for o in orders_today),
            "recent_orders": len(orders),
            "orders_by_agent": orders_by_agent
        }
    except Exception as e:
        print(f"ADMIN STATS ERROR: {e}")
        return {
            "total_agents": len(list_all_agents_admin()),
            "active_agents": len([a for a in list_all_agents_admin() if a.get("active", True)]),
            "orders_today": 0,
            "orders_today_value": 0,
            "recent_orders": 0,
            "orders_by_agent": {}
        }


# ============ Products Routes ============

def filter_items_by_brand(items: List, brands: List[str]) -> List:
    """Filter items to only those matching agent's brands"""
    # Get all brand variations (e.g., "Paper Products Design" -> ["Paper Products Design", "ppd PAPERPRODUCTS DESIGN GmbH", etc.])
    all_patterns = get_all_brand_patterns(brands)
    
    filtered = []
    brand_patterns = [re.compile(re.escape(b), re.IGNORECASE) for b in all_patterns]
    
    for item in items:
        item_brand = item.get("brand") or ""
        item_manufacturer = item.get("manufacturer") or ""
        item_cf_brand = item.get("cf_brand") or ""
        # Also check item group or category if brand field isn't set
        item_group = item.get("group_name") or item.get("category_name") or ""
        
        # Check all text fields against all brand patterns
        all_text = f"{item_brand} {item_manufacturer} {item_cf_brand} {item_group}"
        
        for pattern in brand_patterns:
            if pattern.search(all_text):
                filtered.append(item)
                break
    
    return filtered


@app.get("/api/products/sync")
async def sync_products(
    agent: TokenData = Depends(get_current_agent)
):
    """Get ALL products for offline sync - uses server-side cache to minimize API calls"""
    try:
        print(f"SYNC: Starting product sync for {agent.agent_name}")
        
        # Use cached items - only hits Zoho API every 30 minutes!
        all_zoho_items = await zoho_api.get_all_items_cached()
        
        # Filter to agent's brands (done locally, no API calls)
        all_items = filter_items_by_brand(all_zoho_items, agent.brands)
        
        # Filter out inactive items
        all_items = [item for item in all_items if item.get("status") != "inactive"]
        
        # Sort by SKU alphabetically
        all_items.sort(key=lambda x: (x.get("sku") or "").upper())
        
        # Transform for frontend - include image_url for direct CDN access
        products = []
        for item in all_items:
            sku = item.get("sku", "")
            # Check for custom image URL first (e.g. Elvang Cloudinary), only use valid URLs
            img_url = _image_urls.get(sku) or item.get("image_url")
            if img_url and not img_url.startswith('http'):
                img_url = None
            products.append({
                "item_id": item.get("item_id"),
                "name": item.get("name"),
                "sku": sku,
                "ean": item.get("ean") or item.get("upc") or "",
                "description": item.get("description", ""),
                "rate": item.get("rate", 0),
                "stock_on_hand": item.get("stock_on_hand", 0),
                "brand": item.get("brand") or item.get("manufacturer") or "",
                "unit": item.get("unit", "pcs"),
                "pack_qty": _pack_quantities.get(sku),
                "image_url": img_url
            })
        
        print(f"SYNC: Returning {len(products)} products for {agent.agent_name}")
        
        return {
            "products": products,
            "total": len(products),
            "synced_at": datetime.utcnow().isoformat()
        }
    except Exception as e:
        print(f"SYNC ERROR: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/products/stock")
async def get_stock_levels(
    agent: TokenData = Depends(get_current_agent)
):
    """Get lightweight stock levels only - uses server-side cache"""
    try:
        print(f"STOCK: Getting stock levels for {agent.agent_name}")
        
        # Use cached items - only hits Zoho API every 30 minutes!
        all_zoho_items = await zoho_api.get_all_items_cached()
        
        # Filter to agent's brands (done locally, no API calls)
        agent_items = filter_items_by_brand(all_zoho_items, agent.brands)
        
        # Build stock data
        stock_data = []
        for item in agent_items:
            if item.get("status") != "inactive":
                stock_data.append({
                    "item_id": item.get("item_id"),
                    "sku": item.get("sku"),
                    "stock_on_hand": item.get("stock_on_hand", 0)
                })
        
        print(f"STOCK: Returning {len(stock_data)} stock levels for {agent.agent_name}")
        
        return {
            "stock": stock_data,
            "updated_at": datetime.utcnow().isoformat()
        }
    except Exception as e:
        print(f"STOCK ERROR: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/customers/sync")
async def sync_customers(
    agent: TokenData = Depends(get_current_agent)
):
    """Get ALL customers for offline sync - no pagination"""
    try:
        print(f"SYNC: Starting customer sync for {agent.agent_name}")
        
        all_contacts = []
        page = 1
        
        while True:
            response = await zoho_api.get_contacts(page=page, per_page=200)
            contacts = response.get("contacts", [])
            all_contacts.extend(contacts)
            
            if not response.get("page_context", {}).get("has_more_page", False):
                break
            page += 1
            if page > 50:  # Safety limit
                break
        
        customers = [{
            "contact_id": c.get("contact_id"),
            "company_name": c.get("company_name") or c.get("contact_name"),
            "contact_name": c.get("contact_name"),
            "email": c.get("email"),
            "phone": c.get("phone")
        } for c in all_contacts]
        
        print(f"SYNC: Returning {len(customers)} customers")
        
        return {
            "customers": customers,
            "total": len(customers),
            "synced_at": datetime.utcnow().isoformat()
        }
    except Exception as e:
        print(f"SYNC ERROR: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/products")
async def get_products(
    page: int = 1,
    limit: int = 500,
    search: Optional[str] = None,
    brand: Optional[str] = None,
    agent: TokenData = Depends(get_current_agent)
):
    """Get products filtered by agent's brands - ALWAYS uses server-side cache"""
    try:
        # ALWAYS use cached items - never hit Zoho API directly!
        all_zoho_items = await zoho_api.get_all_items_cached()
        
        # Filter by brand
        filter_brands = [brand] if brand else agent.brands
        items = filter_items_by_brand(all_zoho_items, filter_brands)
        
        # Filter out inactive
        items = [i for i in items if i.get("status") != "inactive"]
        
        # Apply search filter LOCALLY (no API call!)
        if search:
            search_lower = search.lower()
            items = [
                i for i in items
                if search_lower in (i.get("name") or "").lower()
                or search_lower in (i.get("sku") or "").lower()
                or search_lower in (i.get("ean") or i.get("upc") or "").lower()
                or search_lower in (i.get("description") or "").lower()
            ]
        
        # Sort by SKU alphabetically
        items.sort(key=lambda x: (x.get("sku") or "").upper())
        
        # Paginate - use limit param (default 500 to get all)
        per_page = min(limit, 2000)  # Cap at 2000 max
        start = (page - 1) * per_page
        end = start + per_page
        has_more = end < len(items)
        items = items[start:end]
        
        # Transform for frontend (only include selling price, not purchase price)
        # Filter out inactive items
        products = []
        for item in items:
            # Skip inactive items
            if item.get("status") == "inactive":
                continue
            sku = item.get("sku", "")
            # Only return actual URLs, not Zoho document IDs
            img_url = _image_urls.get(sku) or item.get("image_url")
            if img_url and not img_url.startswith('http'):
                img_url = None
            
            products.append({
                "item_id": item.get("item_id"),
                "name": item.get("name"),
                "sku": sku,
                "description": item.get("description", ""),
                "rate": item.get("rate", 0),  # Selling price
                "stock_on_hand": item.get("stock_on_hand", 0),
                "image_url": img_url,
                "brand": item.get("brand") or item.get("manufacturer") or item.get("cf_brand") or item.get("group_name", ""),
                "unit": item.get("unit", "pcs"),
                "status": item.get("status", "active"),
                "pack_qty": _pack_quantities.get(sku)
            })
        
        return {
            "products": products,
            "page": page,
            "has_more": has_more
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/products/{item_id}")
async def get_product(
    item_id: str,
    agent: TokenData = Depends(get_current_agent)
):
    """Get a single product with current stock - uses cache"""
    try:
        # Use cached items - no API call!
        all_items = await zoho_api.get_all_items_cached()
        
        # Find the item
        for item in all_items:
            if item.get("item_id") == item_id:
                return {
                    "item_id": item.get("item_id"),
                    "name": item.get("name"),
                    "sku": item.get("sku"),
                    "description": item.get("description", ""),
                    "rate": item.get("rate", 0),
                    "stock_on_hand": item.get("stock_on_hand", 0),
                    "image_url": item.get("image_url"),
                    "brand": item.get("brand") or item.get("manufacturer") or item.get("cf_brand", ""),
                    "unit": item.get("unit", "pcs")
                }
        
        raise HTTPException(status_code=404, detail="Product not found")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============ Product Images ============

@app.get("/api/products/{item_id}/image")
async def get_product_image(item_id: str):
    """Get product image - uses zoho_api caching"""
    try:
        image_data = await zoho_api.get_item_image(item_id)
        
        if image_data:
            return Response(
                content=image_data, 
                media_type="image/jpeg",
                headers={
                    "Cache-Control": "public, max-age=86400",  # Browser caches for 24 hours
                    "ETag": f'"{item_id}"'
                }
            )
        else:
            raise HTTPException(status_code=404, detail="Image not found")
    except Exception as e:
        raise HTTPException(status_code=404, detail="Image not found")


@app.get("/api/barcode/{barcode}")
async def lookup_barcode(
    barcode: str,
    agent: TokenData = Depends(get_current_agent)
):
    """Look up a product by EAN/barcode - uses cache to avoid API calls"""
    try:
        print(f"BARCODE: Looking up {barcode}")
        
        # Use cached items - no API call!
        all_items = await zoho_api.get_all_items_cached()
        
        # Look for exact match on EAN, UPC, or SKU
        barcode_upper = barcode.upper()
        for item in all_items:
            item_ean = item.get("ean") or item.get("upc") or ""
            item_sku = item.get("sku") or ""
            
            if (item_ean == barcode or item_sku.upper() == barcode_upper):
                if item.get("status") == "inactive":
                    return {"found": False, "message": "Product is inactive"}
                
                # Verify agent has access to this brand
                if not filter_items_by_brand([item], agent.brands):
                    return {"found": False, "message": "Product not available for your brands"}
                
                sku = item.get("sku", "")
                print(f"BARCODE: Found {item.get('name')}")
                return {
                    "found": True,
                    "product": {
                        "item_id": item.get("item_id"),
                        "name": item.get("name"),
                        "sku": sku,
                        "ean": item_ean or barcode,
                        "description": item.get("description", ""),
                        "rate": item.get("rate", 0),
                        "stock_on_hand": item.get("stock_on_hand", 0),
                        "brand": item.get("brand") or item.get("manufacturer") or "",
                        "unit": item.get("unit", "pcs"),
                        "pack_qty": _pack_quantities.get(sku)
                    }
                }
        
        print(f"BARCODE: Not found: {barcode}")
        return {"found": False, "message": "Product not found"}
        
    except Exception as e:
        print(f"BARCODE ERROR: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ Customers Routes ============

@app.get("/api/customers")
async def get_customers(
    page: int = 1,
    search: Optional[str] = None,
    agent: TokenData = Depends(get_current_agent)
):
    """Get customers"""
    try:
        # If searching, fetch multiple pages to search through more customers
        if search:
            all_contacts = []
            for p in range(1, 6):  # Fetch up to 5 pages (1000 customers)
                response = await zoho_api.get_contacts(page=p)
                contacts = response.get("contacts", [])
                all_contacts.extend(contacts)
                if not response.get("page_context", {}).get("has_more_page", False):
                    break
            
            search_lower = search.lower()
            contacts = [
                c for c in all_contacts
                if search_lower in (c.get("company_name") or "").lower()
                or search_lower in (c.get("contact_name") or "").lower()
                or search_lower in (c.get("email") or "").lower()
            ]
            has_more = False
        else:
            response = await zoho_api.get_contacts(page=page)
            contacts = response.get("contacts", [])
            has_more = response.get("page_context", {}).get("has_more_page", False)
        
        customers = []
        for contact in contacts:
            customers.append({
                "contact_id": contact.get("contact_id"),
                "company_name": contact.get("company_name") or contact.get("contact_name"),
                "contact_name": contact.get("contact_name"),
                "email": contact.get("email"),
                "phone": contact.get("phone"),
                "outstanding": contact.get("outstanding_receivable_amount", 0)
            })
        
        return {
            "customers": customers,
            "page": page,
            "has_more": has_more
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/customers")
async def create_customer(
    customer: CustomerCreate,
    agent: TokenData = Depends(get_current_agent)
):
    """Create a new customer"""
    try:
        # Build notes with all the additional info
        notes_parts = [f"Created by {agent.agent_name} via Sales App"]
        if customer.billing_address:
            notes_parts.append(f"\nBilling Address:\n{customer.billing_address}")
        if customer.shipping_address:
            notes_parts.append(f"\nShipping Address:\n{customer.shipping_address}")
        if customer.booking_requirements:
            notes_parts.append(f"\nBooking In Requirements:\n{customer.booking_requirements}")
        if customer.payment_terms:
            notes_parts.append(f"\nPayment Terms: {customer.payment_terms}")
        
        contact_data = {
            "contact_name": customer.company_name,
            "company_name": customer.company_name,
            "contact_type": "customer",
            "notes": "\n".join(notes_parts)
        }
        
        if customer.contact_name:
            contact_data["contact_persons"] = [{
                "first_name": customer.contact_name.split()[0] if customer.contact_name else "",
                "last_name": " ".join(customer.contact_name.split()[1:]) if customer.contact_name and len(customer.contact_name.split()) > 1 else "",
                "email": customer.email,
                "phone": customer.phone,
                "is_primary_contact": True
            }]
        
        if customer.email:
            contact_data["email"] = customer.email
        if customer.phone:
            contact_data["phone"] = customer.phone
        
        # Set billing address as structured data if Zoho supports it
        if customer.billing_address:
            contact_data["billing_address"] = {
                "address": customer.billing_address
            }
        if customer.shipping_address:
            contact_data["shipping_address"] = {
                "address": customer.shipping_address
            }
        
        response = await zoho_api.create_contact(contact_data)
        contact = response.get("contact", {})
        
        return {
            "contact_id": contact.get("contact_id"),
            "company_name": contact.get("company_name"),
            "contact_name": contact.get("contact_name"),
            "message": "Customer created successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============ Orders Routes ============

@app.post("/api/orders")
async def create_order(
    order: OrderCreate,
    agent: TokenData = Depends(get_current_agent)
):
    """Create a new sales order"""
    try:
        # Build line items for Zoho - only include fields Zoho expects
        line_items = []
        for item in order.line_items:
            line_item = {
                "item_id": item.item_id,
                "quantity": item.quantity,
                "rate": item.rate
            }
            if item.discount_percent > 0:
                line_item["discount"] = item.discount_percent
            line_items.append(line_item)
        
        # Build notes with delivery info
        notes_parts = [f"Order placed by {agent.agent_name}"]
        if order.delivery_date:
            notes_parts.append(f"Required delivery date: {order.delivery_date}")
        if order.notes:
            notes_parts.append(order.notes)
        
        order_data = {
            "customer_id": order.customer_id,
            "date": datetime.now().strftime("%Y-%m-%d"),
            "line_items": line_items,
            "notes": "\n".join(notes_parts)
        }
        
        # Add delivery charge as adjustment if applicable
        if order.delivery_charge and order.delivery_charge > 0:
            order_data["adjustment"] = order.delivery_charge
            order_data["adjustment_description"] = "Delivery Charge"
        
        if order.reference_number:
            order_data["reference_number"] = order.reference_number
        
        print(f"ORDER DEBUG: Sending to Zoho: {order_data}")
        
        response = await zoho_api.create_sales_order(order_data)
        salesorder = response.get("salesorder", {})
        
        return {
            "salesorder_id": salesorder.get("salesorder_id"),
            "salesorder_number": salesorder.get("salesorder_number"),
            "total": salesorder.get("total"),
            "status": salesorder.get("status"),
            "message": "Order created successfully"
        }
    except Exception as e:
        error_msg = str(e)
        # Extract Zoho's error message if present
        if "Inactive items" in error_msg:
            error_msg = "Some items in your cart are no longer available. Please clear your cart and try again."
        print(f"ORDER ERROR: {e}")
        raise HTTPException(status_code=500, detail=error_msg)


@app.get("/api/orders")
async def get_orders(
    page: int = 1,
    customer_id: Optional[str] = None,
    agent: TokenData = Depends(get_current_agent)
):
    """Get recent sales orders - filtered by agent unless admin"""
    try:
        # Fetch more orders if we need to filter
        if is_admin(agent.agent_id):
            # Admins see all orders
            response = await zoho_api.get_sales_orders(page=page, customer_id=customer_id)
            orders = response.get("salesorders", [])
            has_more = response.get("page_context", {}).get("has_more_page", False)
        else:
            # Non-admins only see their own orders
            # Need to fetch more and filter since Zoho doesn't filter by notes
            all_orders = []
            current_page = 1
            while len(all_orders) < 200:  # Safety limit
                response = await zoho_api.get_sales_orders(page=current_page, customer_id=customer_id)
                page_orders = response.get("salesorders", [])
                if not page_orders:
                    break
                all_orders.extend(page_orders)
                if not response.get("page_context", {}).get("has_more_page", False):
                    break
                current_page += 1
            
            # Filter to only orders placed by this agent
            # Orders have notes like "Order placed by Kate\n..."
            agent_orders = [
                o for o in all_orders
                if f"Order placed by {agent.agent_name}" in (o.get("notes") or "")
            ]
            
            # Paginate the filtered results
            per_page = 20
            start = (page - 1) * per_page
            end = start + per_page
            orders = agent_orders[start:end]
            has_more = end < len(agent_orders)
        
        return {
            "orders": [{
                "salesorder_id": o.get("salesorder_id"),
                "salesorder_number": o.get("salesorder_number"),
                "customer_name": o.get("customer_name"),
                "date": o.get("date"),
                "total": o.get("total"),
                "status": o.get("status")
            } for o in orders],
            "page": page,
            "has_more": has_more
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/orders/{salesorder_id}")
async def get_order(
    salesorder_id: str,
    agent: TokenData = Depends(get_current_agent)
):
    """Get a single order with full details - only if agent has permission"""
    try:
        response = await zoho_api.get_sales_order(salesorder_id)
        order = response.get("salesorder", {})
        
        # Check permission - admins can see all, others only their own
        if not is_admin(agent.agent_id):
            notes = order.get("notes") or ""
            if f"Order placed by {agent.agent_name}" not in notes:
                raise HTTPException(status_code=403, detail="You don't have permission to view this order")
        
        return order
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============ Export Quote ============

class ExportItem(BaseModel):
    item_id: str
    name: str
    sku: str
    ean: Optional[str] = None
    rate: float
    quantity: int

class ExportRequest(BaseModel):
    items: List[ExportItem]
    customer_name: Optional[str] = None
    include_images: bool = False  # Default OFF to save API calls

@app.post("/api/export/quote")
async def export_quote(
    request: ExportRequest,
    agent: TokenData = Depends(get_current_agent)
):
    """Export cart as Excel quote - images optional to save API calls"""
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from PIL import Image
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Quote"
        
        # Set up header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column widths - adjust if no images
        if request.include_images:
            ws.column_dimensions['A'].width = 15  # Image
            ws.column_dimensions['B'].width = 40  # Description
            ws.column_dimensions['C'].width = 15  # SKU
            ws.column_dimensions['D'].width = 18  # EAN
            ws.column_dimensions['E'].width = 10  # Qty
            ws.column_dimensions['F'].width = 12  # Price
            ws.column_dimensions['G'].width = 12  # Total
            headers = ['Image', 'Description', 'SKU', 'EAN', 'Qty', 'Unit Price', 'Total']
        else:
            ws.column_dimensions['A'].width = 40  # Description
            ws.column_dimensions['B'].width = 15  # SKU
            ws.column_dimensions['C'].width = 18  # EAN
            ws.column_dimensions['D'].width = 10  # Qty
            ws.column_dimensions['E'].width = 12  # Price
            ws.column_dimensions['F'].width = 12  # Total
            headers = ['Description', 'SKU', 'EAN', 'Qty', 'Unit Price', 'Total']
        
        # Add title
        if request.customer_name:
            end_col = 'G' if request.include_images else 'F'
            ws.merge_cells(f'A1:{end_col}1')
            ws['A1'] = f"Quote for {request.customer_name}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal="center")
            start_row = 3
        else:
            start_row = 1
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        ws.row_dimensions[start_row].height = 25
        
        # Add items
        current_row = start_row + 1
        grand_total = 0
        
        for item in request.items:
            line_total = item.rate * item.quantity
            grand_total += line_total
            
            if request.include_images:
                # Set row height for image
                ws.row_dimensions[current_row].height = 75
                
                # Try to fetch and add image
                try:
                    image_data = await zoho_api.get_item_image(item.item_id)
                    if image_data:
                        img = Image.open(io.BytesIO(image_data))
                        img.thumbnail((90, 90), Image.Resampling.LANCZOS)
                        img_bytes = io.BytesIO()
                        img.save(img_bytes, format='PNG')
                        img_bytes.seek(0)
                        xl_img = XLImage(img_bytes)
                        xl_img.width = 90
                        xl_img.height = 90
                        ws.add_image(xl_img, f'A{current_row}')
                except Exception as img_err:
                    print(f"IMAGE EXPORT ERROR: {img_err}")
                
                # Add item details with image column offset
                ws.cell(row=current_row, column=2, value=item.name).border = thin_border
                ws.cell(row=current_row, column=3, value=item.sku).border = thin_border
                ws.cell(row=current_row, column=4, value=item.ean or '').border = thin_border
                ws.cell(row=current_row, column=5, value=item.quantity).border = thin_border
                ws.cell(row=current_row, column=5).alignment = Alignment(horizontal="center")
                price_cell = ws.cell(row=current_row, column=6, value=item.rate)
                price_cell.number_format = '£#,##0.00'
                price_cell.border = thin_border
                total_cell = ws.cell(row=current_row, column=7, value=line_total)
                total_cell.number_format = '£#,##0.00'
                total_cell.border = thin_border
                for col in [3, 4]:
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal="center")
            else:
                # No images - simpler layout
                ws.cell(row=current_row, column=1, value=item.name).border = thin_border
                ws.cell(row=current_row, column=2, value=item.sku).border = thin_border
                ws.cell(row=current_row, column=3, value=item.ean or '').border = thin_border
                ws.cell(row=current_row, column=4, value=item.quantity).border = thin_border
                ws.cell(row=current_row, column=4).alignment = Alignment(horizontal="center")
                price_cell = ws.cell(row=current_row, column=5, value=item.rate)
                price_cell.number_format = '£#,##0.00'
                price_cell.border = thin_border
                total_cell = ws.cell(row=current_row, column=6, value=line_total)
                total_cell.number_format = '£#,##0.00'
                total_cell.border = thin_border
                for col in [2, 3]:
                    ws.cell(row=current_row, column=col).alignment = Alignment(horizontal="center")
            
            current_row += 1
        
        # Add grand total
        current_row += 1
        total_col = 6 if request.include_images else 5
        value_col = 7 if request.include_images else 6
        ws.cell(row=current_row, column=total_col, value="TOTAL:").font = Font(bold=True)
        total_cell = ws.cell(row=current_row, column=value_col, value=grand_total)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '£#,##0.00'
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        date_str = datetime.now().strftime("%Y%m%d")
        customer_part = request.customer_name.replace(' ', '_')[:20] if request.customer_name else 'Quote'
        filename = f"{customer_part}_{date_str}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        print(f"EXPORT ERROR: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============ Quote PDF Export ============

class QuotePDFItem(BaseModel):
    item_id: str
    name: str
    sku: str
    ean: Optional[str] = None
    rate: float
    quantity: int
    discount: float = 0

class QuotePDFRequest(BaseModel):
    items: List[QuotePDFItem]
    customer_name: Optional[str] = None
    customer_email: Optional[str] = None
    include_images: bool = True
    doc_type: str = "quote"  # "quote" or "order"

@app.post("/api/export/quote-pdf")
async def export_quote_pdf(
    request: QuotePDFRequest,
    agent: TokenData = Depends(get_current_agent)
):
    """Generate a PDF quote from cart items with product images"""
    try:
        import httpx
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, KeepTogether
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        from PIL import Image
        
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4, 
            topMargin=15*mm, 
            bottomMargin=15*mm,
            leftMargin=12*mm,
            rightMargin=12*mm
        )
        
        # Page dimensions
        page_width = A4[0] - 24*mm  # Width minus margins
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=20, alignment=TA_CENTER, spaceAfter=2*mm)
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, textColor=colors.grey)
        header_style = ParagraphStyle('Header', parent=styles['Heading2'], fontSize=14, spaceAfter=4*mm)
        normal_style = styles['Normal']
        cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontSize=9, leading=11)
        cell_style_small = ParagraphStyle('CellSmall', parent=styles['Normal'], fontSize=8, leading=10, textColor=colors.grey)
        
        elements = []
        
        # Header - dynamic based on doc_type
        doc_title = "Order Confirmation" if request.doc_type == "order" else "Product Quotation"
        elements.append(Paragraph("DM Brands Ltd", title_style))
        elements.append(Paragraph(doc_title, subtitle_style))
        elements.append(Spacer(1, 6*mm))
        
        # Quote info
        date_str = datetime.now().strftime("%d %B %Y")
        info_data = []
        if request.customer_name:
            info_data.append([Paragraph("<b>Customer:</b>", normal_style), Paragraph(request.customer_name, normal_style)])
        info_data.append([Paragraph("<b>Date:</b>", normal_style), Paragraph(date_str, normal_style)])
        info_data.append([Paragraph("<b>Prepared by:</b>", normal_style), Paragraph(agent.agent_name, normal_style)])
        if request.doc_type == "quote":
            info_data.append([Paragraph("<b>Valid for:</b>", normal_style), Paragraph("30 days", normal_style)])
        
        if info_data:
            info_table = Table(info_data, colWidths=[70, page_width - 70])
            info_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 6*mm))
        
        # Fetch images from Cloudinary if requested
        image_cache = {}
        if request.include_images:
            cloudinary_base = f"https://res.cloudinary.com/{settings.cloudinary_cloud_name}/image/upload"
            async with httpx.AsyncClient(timeout=10.0) as client:
                for item in request.items:
                    if item.sku:
                        # Convert SKU for Cloudinary (dots to underscores for My Flame)
                        cloudinary_sku = item.sku.replace('.', '_')
                        img_url = f"{cloudinary_base}/w_120,h_120,c_pad,b_white,q_80/products/{cloudinary_sku}.jpg"
                        try:
                            response = await client.get(img_url)
                            if response.status_code == 200:
                                image_cache[item.sku] = response.content
                        except:
                            pass  # Skip failed images
        
        # Build product rows - each product is a mini-table row
        # Column widths: Image (25mm), Details (flex), Qty (18mm), Price (22mm), Total (25mm)
        if request.include_images:
            col_widths = [25*mm, page_width - 25*mm - 18*mm - 22*mm - 25*mm, 18*mm, 22*mm, 25*mm]
        else:
            col_widths = [page_width - 18*mm - 22*mm - 25*mm, 18*mm, 22*mm, 25*mm]
        
        # Header row
        header_fill = colors.HexColor('#1e3a5f')
        header_text = colors.white
        
        if request.include_images:
            header_row = [
                Paragraph("<font color='white'><b>Image</b></font>", cell_style),
                Paragraph("<font color='white'><b>Product Details</b></font>", cell_style),
                Paragraph("<font color='white'><b>Qty</b></font>", cell_style),
                Paragraph("<font color='white'><b>Price</b></font>", cell_style),
                Paragraph("<font color='white'><b>Total</b></font>", cell_style),
            ]
        else:
            header_row = [
                Paragraph("<font color='white'><b>Product Details</b></font>", cell_style),
                Paragraph("<font color='white'><b>Qty</b></font>", cell_style),
                Paragraph("<font color='white'><b>Price</b></font>", cell_style),
                Paragraph("<font color='white'><b>Total</b></font>", cell_style),
            ]
        
        table_data = [header_row]
        grand_total = 0
        
        for item in request.items:
            line_total = item.rate * item.quantity
            if item.discount > 0:
                line_total = line_total * (1 - item.discount / 100)
            grand_total += line_total
            
            # Product details cell
            details_parts = [f"<b>{item.name}</b>"]
            details_parts.append(f"<font size='8' color='grey'>SKU: {item.sku}</font>")
            if item.ean:
                details_parts.append(f"<font size='8' color='grey'>EAN: {item.ean}</font>")
            if item.discount > 0:
                details_parts.append(f"<font size='8' color='#c00'>Discount: {item.discount:.0f}%</font>")
            
            details_cell = Paragraph("<br/>".join(details_parts), cell_style)
            
            # Price display
            price_text = f"£{item.rate:.2f}"
            total_text = f"£{line_total:.2f}"
            
            if request.include_images:
                # Image cell
                if item.sku in image_cache:
                    try:
                        img_data = io.BytesIO(image_cache[item.sku])
                        img = RLImage(img_data, width=22*mm, height=22*mm)
                        img_cell = img
                    except:
                        img_cell = ""
                else:
                    img_cell = ""
                
                row = [
                    img_cell,
                    details_cell,
                    Paragraph(str(item.quantity), cell_style),
                    Paragraph(price_text, cell_style),
                    Paragraph(f"<b>{total_text}</b>", cell_style),
                ]
            else:
                row = [
                    details_cell,
                    Paragraph(str(item.quantity), cell_style),
                    Paragraph(price_text, cell_style),
                    Paragraph(f"<b>{total_text}</b>", cell_style),
                ]
            
            table_data.append(row)
        
        # Create main table
        main_table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # Table styling
        num_rows = len(table_data)
        style_commands = [
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), header_fill),
            ('TEXTCOLOR', (0, 0), (-1, 0), header_text),
            
            # All cells
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            
            # Alignment
            ('ALIGN', (-3, 1), (-3, -1), 'CENTER'),  # Qty
            ('ALIGN', (-2, 0), (-1, -1), 'RIGHT'),   # Prices
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e0')),
            ('LINEBELOW', (0, 0), (-1, 0), 1, header_fill),
        ]
        
        # Alternating row colors
        for i in range(2, num_rows, 2):
            style_commands.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8f9fa')))
        
        main_table.setStyle(TableStyle(style_commands))
        elements.append(main_table)
        
        # Totals section
        elements.append(Spacer(1, 4*mm))
        
        total_col_offset = 3 if request.include_images else 2
        totals_data = [
            ["", "", Paragraph("<b>Subtotal (ex VAT):</b>", cell_style), Paragraph(f"<b>£{grand_total:.2f}</b>", cell_style)],
        ]
        
        totals_table = Table(totals_data, colWidths=[page_width - 70 - 80, 10, 70, 80])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (-2, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(totals_table)
        
        # Footer
        elements.append(Spacer(1, 10*mm))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=9, textColor=colors.grey, alignment=TA_CENTER)
        elements.append(Paragraph("All prices exclude VAT. E&OE.", footer_style))
        elements.append(Paragraph("DM Brands Ltd | sales@dmbrands.co.uk | www.dmbrands.co.uk", footer_style))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Generate filename
        date_str = datetime.now().strftime("%Y%m%d")
        customer_part = request.customer_name.replace(' ', '_')[:20] if request.customer_name else 'Customer'
        doc_prefix = "Order" if request.doc_type == "order" else "Quote"
        filename = f"{doc_prefix}_{customer_part}_{date_str}.pdf"
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "X-Filename": filename  # For frontend to read
            }
        )
        
    except Exception as e:
        print(f"QUOTE PDF ERROR: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ============ Order PDF Export ============

@app.get("/api/orders/{salesorder_id}/pdf")
async def export_order_pdf(
    salesorder_id: str,
    agent: TokenData = Depends(get_current_agent)
):
    """Generate a PDF of a sales order"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        
        # Fetch order from Zoho
        response = await zoho_api.get_sales_order(salesorder_id)
        order = response.get("salesorder", {})
        
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        # Check permission - admins can see all, others only their own
        if not is_admin(agent.agent_id):
            notes = order.get("notes") or ""
            if f"Order placed by {agent.agent_name}" not in notes:
                raise HTTPException(status_code=403, detail="You don't have permission to view this order")
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4, 
            topMargin=15*mm, 
            bottomMargin=15*mm,
            leftMargin=15*mm,
            rightMargin=15*mm
        )
        
        # Page width for calculating column widths (A4 = 210mm, minus margins)
        page_width = 180*mm
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER)
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=14, alignment=TA_CENTER)
        normal_style = styles['Normal']
        
        # Cell styles for proper text wrapping
        cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=8, leading=10)
        cell_style_bold = ParagraphStyle('CellStyleBold', parent=cell_style, fontName='Helvetica-Bold')
        
        elements = []
        
        # Header
        elements.append(Paragraph("DM Brands Ltd", title_style))
        elements.append(Paragraph("Sales Order Confirmation", subtitle_style))
        elements.append(Spacer(1, 8*mm))
        
        # Order details
        order_info = [
            [Paragraph("<b>Order Number:</b>", normal_style), Paragraph(order.get("salesorder_number", ""), normal_style)],
            [Paragraph("<b>Date:</b>", normal_style), Paragraph(order.get("date", ""), normal_style)],
            [Paragraph("<b>Customer:</b>", normal_style), Paragraph(order.get("customer_name", ""), normal_style)],
            [Paragraph("<b>Status:</b>", normal_style), Paragraph(order.get("status", "").title(), normal_style)],
        ]
        
        info_table = Table(order_info, colWidths=[80, page_width - 80])
        info_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 8*mm))
        
        # Line items table - column widths: SKU(50), Description(flex), Qty(30), Price(50), Total(50)
        col_widths = [50, page_width - 50 - 30 - 50 - 50, 30, 50, 50]
        
        # Header row using Paragraphs
        header_style_white = ParagraphStyle('HeaderWhite', parent=cell_style, textColor=colors.white, alignment=TA_CENTER)
        table_data = [[
            Paragraph("<b>SKU</b>", header_style_white),
            Paragraph("<b>Description</b>", header_style_white),
            Paragraph("<b>Qty</b>", header_style_white),
            Paragraph("<b>Price</b>", header_style_white),
            Paragraph("<b>Total</b>", header_style_white),
        ]]
        
        # Line items with Paragraph objects for text wrapping
        line_items = order.get("line_items", [])
        for item in line_items:
            table_data.append([
                Paragraph(item.get("sku", ""), cell_style),
                Paragraph(item.get("name", ""), cell_style),  # Will wrap automatically
                Paragraph(str(item.get("quantity", 0)), cell_style),
                Paragraph(f"£{item.get('rate', 0):.2f}", cell_style),
                Paragraph(f"£{item.get('item_total', 0):.2f}", cell_style),
            ])
        
        # Add totals
        table_data.append(["", "", "", Paragraph("<b>Subtotal:</b>", cell_style), Paragraph(f"£{order.get('sub_total', 0):.2f}", cell_style)])
        if order.get("adjustment"):
            table_data.append(["", "", "", Paragraph("<b>Delivery:</b>", cell_style), Paragraph(f"£{order.get('adjustment', 0):.2f}", cell_style)])
        table_data.append(["", "", "", Paragraph("<b>Total (ex VAT):</b>", cell_style_bold), Paragraph(f"<b>£{order.get('total', 0):.2f}</b>", cell_style)])
        
        items_table = Table(table_data, colWidths=col_widths)
        
        num_items = len(line_items)
        items_table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
            
            # All cells
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            
            # Alignment
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),  # Qty
            ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),  # Prices
            
            # Grid for header + line items only
            ('GRID', (0, 0), (-1, num_items), 0.5, colors.grey),
            
            # Alternating row colors
            *[('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8fafc')) for i in range(2, num_items + 1, 2)],
            
            # Line above totals
            ('LINEABOVE', (3, num_items + 1), (-1, num_items + 1), 1, colors.grey),
        ]))
        elements.append(items_table)
        elements.append(Spacer(1, 8*mm))
        
        # Notes
        if order.get("notes"):
            notes_style = ParagraphStyle('NotesStyle', parent=normal_style, fontSize=9, leading=12)
            elements.append(Paragraph("<b>Notes:</b>", normal_style))
            elements.append(Spacer(1, 2*mm))
            elements.append(Paragraph(order.get("notes", "").replace("\n", "<br/>"), notes_style))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Generate filename
        filename = f"{order.get('salesorder_number', 'order')}.pdf"
        
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"PDF EXPORT ERROR: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ============ Health Check ============

@app.get("/api/debug/image-info")
async def debug_image_info():
    """Debug: Check what image data Zoho provides"""
    try:
        # Get a few items from cache
        all_items = await zoho_api.get_all_items_cached()
        
        # Find items with images
        items_with_images = []
        for item in all_items[:50]:  # Check first 50
            if item.get("image_document_id") or item.get("image_url") or item.get("image_name"):
                items_with_images.append({
                    "name": item.get("name"),
                    "sku": item.get("sku"),
                    "image_document_id": item.get("image_document_id"),
                    "image_url": item.get("image_url"),
                    "image_name": item.get("image_name"),
                    "all_image_fields": {k: v for k, v in item.items() if 'image' in k.lower()}
                })
        
        return {
            "total_items_checked": 50,
            "items_with_image_data": len(items_with_images),
            "samples": items_with_images[:5]
        }
    except Exception as e:
        import traceback
        return {"error": str(e), "traceback": traceback.format_exc()}


# ============ Catalogues (Database) ============

from database import SessionLocal, Catalogue as CatalogueModel

def load_catalogues():
    """Load all catalogues from database"""
    db = SessionLocal()
    try:
        catalogues = db.query(CatalogueModel).all()
        return [
            {
                "id": cat.id,
                "brand": cat.brand,
                "name": cat.name,
                "description": cat.description or "",
                "url": cat.url,
                "size_mb": cat.size_mb or 0,
                "updated": cat.updated_at.strftime("%Y-%m-%d") if cat.updated_at else "",
                "added_by": cat.added_by or ""
            }
            for cat in catalogues
        ]
    finally:
        db.close()


def save_catalogue(catalogue_data: dict):
    """Save a catalogue to database"""
    db = SessionLocal()
    try:
        cat = db.query(CatalogueModel).filter(CatalogueModel.id == catalogue_data["id"]).first()
        if cat:
            # Update existing
            cat.brand = catalogue_data.get("brand", cat.brand)
            cat.name = catalogue_data.get("name", cat.name)
            cat.description = catalogue_data.get("description", cat.description)
            cat.url = catalogue_data.get("url", cat.url)
            cat.size_mb = catalogue_data.get("size_mb", cat.size_mb)
            cat.added_by = catalogue_data.get("added_by", cat.added_by)
        else:
            # Create new
            cat = CatalogueModel(
                id=catalogue_data["id"],
                brand=catalogue_data["brand"],
                name=catalogue_data["name"],
                description=catalogue_data.get("description", ""),
                url=catalogue_data["url"],
                size_mb=catalogue_data.get("size_mb", 0),
                added_by=catalogue_data.get("added_by", "")
            )
            db.add(cat)
        db.commit()
        return cat.id
    finally:
        db.close()


def delete_catalogue_from_db(catalogue_id: str) -> bool:
    """Delete a catalogue from database"""
    db = SessionLocal()
    try:
        cat = db.query(CatalogueModel).filter(CatalogueModel.id == catalogue_id).first()
        if cat:
            db.delete(cat)
            db.commit()
            return True
        return False
    finally:
        db.close()


@app.get("/api/catalogues")
async def get_catalogues(agent: TokenData = Depends(get_current_agent)):
    """Get available catalogues - filtered by agent's brands"""
    catalogues = load_catalogues()
    
    # Filter to agent's brands using the brand patterns
    agent_patterns = get_all_brand_patterns(agent.brands)
    
    filtered = []
    for cat in catalogues:
        # Only include catalogues that have a URL set
        if not cat.get("url"):
            continue
        cat_brand = cat.get("brand", "")
        for pattern in agent_patterns:
            if pattern.lower() in cat_brand.lower() or cat_brand.lower() in pattern.lower():
                filtered.append(cat)
                break
    
    return {"catalogues": filtered}


@app.get("/api/admin/catalogues")
async def admin_list_catalogues(agent: TokenData = Depends(require_admin)):
    """List all catalogues with admin details"""
    catalogues = load_catalogues()
    return {
        "catalogues": catalogues,
        "available_brands": get_all_brands()
    }


class CatalogueCreate(BaseModel):
    brand: str
    name: str
    description: str = ""
    url: str
    size_mb: float = 0


@app.post("/api/admin/catalogues")
async def admin_add_catalogue(
    catalogue: CatalogueCreate,
    agent: TokenData = Depends(require_admin)
):
    """Add a new catalogue with external URL (admin only)"""
    # Generate ID
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    safe_brand = catalogue.brand.lower().replace(" ", "-").replace("ä", "a").replace("ü", "u").replace("ö", "o")
    catalogue_id = f"{safe_brand}-{timestamp}"
    
    new_catalogue = {
        "id": catalogue_id,
        "brand": catalogue.brand,
        "name": catalogue.name,
        "description": catalogue.description,
        "url": catalogue.url,
        "size_mb": catalogue.size_mb,
        "added_by": agent.agent_name
    }
    
    save_catalogue(new_catalogue)
    
    return {
        "message": "Catalogue added successfully",
        "catalogue": new_catalogue
    }


@app.put("/api/admin/catalogues/{catalogue_id}")
async def admin_update_catalogue(
    catalogue_id: str,
    updates: CatalogueCreate,
    agent: TokenData = Depends(require_admin)
):
    """Update a catalogue (admin only)"""
    catalogue_data = {
        "id": catalogue_id,
        "brand": updates.brand,
        "name": updates.name,
        "description": updates.description,
        "url": updates.url,
        "size_mb": updates.size_mb
    }
    
    # Check if exists first
    catalogues = load_catalogues()
    if not any(cat["id"] == catalogue_id for cat in catalogues):
        raise HTTPException(status_code=404, detail="Catalogue not found")
    
    save_catalogue(catalogue_data)
    return {"message": "Catalogue updated successfully"}


@app.delete("/api/admin/catalogues/{catalogue_id}")
async def admin_delete_catalogue(
    catalogue_id: str,
    agent: TokenData = Depends(require_admin)
):
    """Delete a catalogue (admin only)"""
    if not delete_catalogue_from_db(catalogue_id):
        raise HTTPException(status_code=404, detail="Catalogue not found")
    
    return {"message": "Catalogue deleted successfully"}


# ============ Static Product Feed ============

from database import SessionLocal, ProductFeed


@app.get("/api/feed/products")
async def get_product_feed():
    """
    Get the static product feed JSON from database.
    This is the main endpoint for fast sync - returns pre-generated product data.
    """
    db = SessionLocal()
    try:
        feed = db.query(ProductFeed).filter(ProductFeed.id == "main").first()
        
        if feed and feed.feed_json:
            # Parse the stored JSON
            return json.loads(feed.feed_json)
        else:
            # Feed not generated yet
            return {
                "generated_at": None,
                "total_products": 0,
                "products": [],
                "message": "Feed not yet generated. Run /api/cron/generate-feed to create it."
            }
    finally:
        db.close()


@app.get("/api/feed/products-url")
async def get_product_feed_url():
    """Get URL for static product feed"""
    return {
        "url": "/api/feed/products",
        "description": "Static product feed, updated every 4 hours"
    }


@app.post("/api/admin/generate-feed")
async def admin_generate_feed(agent: TokenData = Depends(require_admin)):
    """Manually trigger product feed generation (admin only)"""
    import subprocess
    import sys
    
    script_path = os.path.join(os.path.dirname(__file__), "scripts", "generate_product_feed.py")
    
    if not os.path.exists(script_path):
        raise HTTPException(status_code=500, detail="Feed generator script not found")
    
    try:
        # Run the script
        result = subprocess.run(
            [sys.executable, script_path],
            capture_output=True,
            text=True,
            timeout=300,  # 5 min timeout
            cwd=os.path.dirname(script_path)
        )
        
        if result.returncode != 0:
            return {
                "success": False,
                "error": result.stderr,
                "output": result.stdout
            }
        
        return {
            "success": True,
            "message": "Feed generated successfully",
            "output": result.stdout[-2000:] if len(result.stdout) > 2000 else result.stdout
        }
    except subprocess.TimeoutExpired:
        raise HTTPException(status_code=500, detail="Feed generation timed out")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/cron/generate-feed")
async def cron_generate_feed(secret: str = None):
    """
    Cron endpoint to generate product feed.
    Protected by CRON_SECRET environment variable.
    
    Call with: POST /api/cron/generate-feed?secret=YOUR_CRON_SECRET
    
    Returns immediately and runs feed generation in background.
    """
    # Verify cron secret
    if not settings.cron_secret:
        raise HTTPException(status_code=500, detail="CRON_SECRET not configured")
    
    if secret != settings.cron_secret:
        raise HTTPException(status_code=401, detail="Invalid cron secret")
    
    import subprocess
    import sys
    
    script_path = os.path.join(os.path.dirname(__file__), "scripts", "generate_product_feed.py")
    
    if not os.path.exists(script_path):
        raise HTTPException(status_code=500, detail="Feed generator script not found")
    
    try:
        # Fire and forget - start process in background and return immediately
        # This allows cron services with short timeouts (like cron-job.org's 30s) to work
        subprocess.Popen(
            [sys.executable, script_path],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            cwd=os.path.dirname(script_path),
            start_new_session=True  # Detach from parent process
        )
        
        return {
            "success": True,
            "message": "Feed generation started in background",
            "timestamp": datetime.utcnow().isoformat()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============ Health Check ============

@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "timestamp": datetime.utcnow().isoformat()}


@app.get("/api/debug/pack-qty/{sku}")
async def debug_pack_qty(sku: str):
    """Debug: Check pack quantity for a SKU"""
    return {
        "sku": sku,
        "pack_qty": _pack_quantities.get(sku),
        "total_pack_qtys_loaded": len(_pack_quantities),
        "sample_keys": list(_pack_quantities.keys())[:10]
    }


# ============ Image Manifest (Cloudinary) ============

# Cache for image manifest - stores which SKUs have multiple images
_image_manifest_cache = {
    "data": {},  # SKU -> list of image suffixes e.g. {"SKU123": ["", "_1", "_2"], "SKU456": [""]}
    "updated_at": None
}
_IMAGE_MANIFEST_CACHE_HOURS = 6  # Refresh every 6 hours


@app.get("/api/images/manifest")
async def get_image_manifest():
    """
    Get manifest of all product images in Cloudinary.
    Returns SKU -> list of available image suffixes.
    Cached for 6 hours to avoid hammering Cloudinary API.
    """
    import httpx
    
    # Check cache
    if _image_manifest_cache["updated_at"]:
        cache_age = datetime.utcnow() - _image_manifest_cache["updated_at"]
        if cache_age < timedelta(hours=_IMAGE_MANIFEST_CACHE_HOURS):
            return {
                "manifest": _image_manifest_cache["data"],
                "cached_at": _image_manifest_cache["updated_at"].isoformat(),
                "from_cache": True
            }
    
    # Need to refresh - query Cloudinary Admin API
    if not settings.cloudinary_cloud_name or not settings.cloudinary_api_key:
        return {"manifest": {}, "error": "Cloudinary not configured"}
    
    try:
        manifest = {}
        next_cursor = None
        
        async with httpx.AsyncClient(timeout=30.0) as client:
            # Paginate through all resources in products folder
            while True:
                url = f"https://api.cloudinary.com/v1_1/{settings.cloudinary_cloud_name}/resources/image"
                params = {
                    "type": "upload",
                    "prefix": "products/",
                    "max_results": 500
                }
                if next_cursor:
                    params["next_cursor"] = next_cursor
                
                response = await client.get(
                    url,
                    params=params,
                    auth=(settings.cloudinary_api_key, settings.cloudinary_api_secret)
                )
                
                if response.status_code != 200:
                    print(f"Cloudinary API error: {response.status_code} - {response.text}")
                    break
                
                data = response.json()
                resources = data.get("resources", [])
                
                for resource in resources:
                    # public_id is like "products/SKU123" or "products/SKU123_1"
                    public_id = resource.get("public_id", "")
                    if public_id.startswith("products/"):
                        filename = public_id[9:]  # Remove "products/" prefix
                        
                        # Parse SKU and suffix
                        # e.g. "SKU123" -> SKU="SKU123", suffix=""
                        # e.g. "SKU123_1" -> SKU="SKU123", suffix="_1"
                        if "_" in filename and filename.split("_")[-1].isdigit():
                            parts = filename.rsplit("_", 1)
                            sku = parts[0]
                            suffix = f"_{parts[1]}"
                        else:
                            sku = filename
                            suffix = ""
                        
                        if sku not in manifest:
                            manifest[sku] = []
                        if suffix not in manifest[sku]:
                            manifest[sku].append(suffix)
                
                # Check for more pages
                next_cursor = data.get("next_cursor")
                if not next_cursor:
                    break
        
        # Sort suffixes for each SKU
        for sku in manifest:
            manifest[sku].sort()
        
        # Update cache
        _image_manifest_cache["data"] = manifest
        _image_manifest_cache["updated_at"] = datetime.utcnow()
        
        return {
            "manifest": manifest,
            "total_skus": len(manifest),
            "skus_with_multiple": len([s for s in manifest if len(manifest[s]) > 1]),
            "cached_at": _image_manifest_cache["updated_at"].isoformat(),
            "from_cache": False
        }
        
    except Exception as e:
        print(f"IMAGE MANIFEST ERROR: {e}")
        import traceback
        traceback.print_exc()
        return {"manifest": {}, "error": str(e)}


@app.post("/api/admin/refresh-image-manifest")
async def admin_refresh_image_manifest(agent: TokenData = Depends(require_admin)):
    """Force refresh of the image manifest cache (admin only)"""
    _image_manifest_cache["updated_at"] = None  # Clear cache
    result = await get_image_manifest()
    return {
        "message": "Image manifest refreshed",
        "total_skus": result.get("total_skus", 0),
        "skus_with_multiple": result.get("skus_with_multiple", 0)
    }


# ============ German Stock ============

# Load German stock data
GERMAN_STOCK_FILES = {
    "raeder": os.path.join(os.path.dirname(__file__), "german_stock_raeder.json")
}

def load_german_stock(brand: str) -> dict:
    """Load German stock data for a brand"""
    file_path = GERMAN_STOCK_FILES.get(brand.lower())
    if file_path and os.path.exists(file_path):
        with open(file_path, "r") as f:
            return json.load(f)
    return {}

# Pre-load German stock at startup
_german_stock_cache = {}
for brand, path in GERMAN_STOCK_FILES.items():
    if os.path.exists(path):
        with open(path, "r") as f:
            _german_stock_cache[brand] = json.load(f)
            print(f"STARTUP: Loaded German stock for {brand}: {_german_stock_cache[brand].get('item_count', 0)} items")


@app.get("/api/german-stock/{brand}")
async def get_german_stock(brand: str):
    """Get German warehouse stock data for a brand"""
    brand_lower = brand.lower()
    if brand_lower not in _german_stock_cache:
        raise HTTPException(status_code=404, detail=f"No German stock data for brand: {brand}")
    
    data = _german_stock_cache[brand_lower]
    return {
        "brand": data.get("brand"),
        "updated": data.get("updated"),
        "item_count": data.get("item_count"),
        "items": data.get("items", {})
    }


@app.get("/api/german-stock")
async def get_all_german_stock():
    """Get all German warehouse stock data"""
    return {
        "brands": list(_german_stock_cache.keys()),
        "data": _german_stock_cache
    }


# ============ Static Files (Production) ============

# Serve frontend static files in production
static_dir = os.path.join(os.path.dirname(__file__), "static")
print(f"STARTUP: Looking for static dir at: {static_dir}")
print(f"STARTUP: Static dir exists: {os.path.exists(static_dir)}")

if os.path.exists(static_dir):
    print(f"STARTUP: Static dir contents: {os.listdir(static_dir)}")
    assets_dir = os.path.join(static_dir, "assets")
    
    # Mount static assets (js, css, images) if they exist
    if os.path.exists(assets_dir):
        app.mount("/assets", StaticFiles(directory=assets_dir), name="assets")
        print(f"STARTUP: Mounted /assets from {assets_dir}")
    
    # Explicit root route - serves index.html
    @app.get("/")
    async def serve_root():
        return FileResponse(os.path.join(static_dir, "index.html"))
    
    # Serve index.html for all non-API routes (SPA routing)
    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        # Don't intercept API routes
        if full_path.startswith("api/"):
            raise HTTPException(status_code=404)
        
        # Serve static files if they exist
        file_path = os.path.join(static_dir, full_path)
        if os.path.isfile(file_path):
            return FileResponse(file_path)
        
        # Otherwise serve index.html for SPA routing
        return FileResponse(os.path.join(static_dir, "index.html"))
else:
    print("STARTUP: WARNING - No static directory found! Frontend will not be served.")
    print(f"STARTUP: Current working directory: {os.getcwd()}")
    print(f"STARTUP: __file__ directory: {os.path.dirname(__file__)}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
